"""需求问卷生成器 — 从 questionnaire.md 生成 PDF + XLSX 审查表。

用法:
    python gen_questionnaire.py questionnaire.md        # 生成 PDF + XLSX
    python gen_questionnaire.py questionnaire.md --pdf  # 仅 PDF

依赖: reportlab（PDF）+ openpyxl（XLSX）
"""
from __future__ import annotations

import re
import sys
from pathlib import Path


def parse_questions(md_path: Path) -> list[dict]:
    """解析问卷 md → 问题列表 [{chapter, q, importance, purpose}]。"""
    questions = []
    chapter = "未分类"
    for line in md_path.read_text(encoding="utf-8").splitlines():
        if line.startswith("## "):
            chapter = line[3:].strip()
        m = re.match(r"^\|\s*(\d+)\s*\|\s*([^|]+?)\s*\|\s*(★+)\s*\|\s*([^|]*?)\s*\|", line)
        if m:
            questions.append({
                "num": int(m.group(1)), "q": m.group(2).strip(),
                "importance": m.group(3), "purpose": m.group(4).strip(),
                "chapter": chapter,
            })
    return questions


def gen_pdf(questions: list[dict], out: Path):
    """用 reportlab 生成中文问卷 PDF（SimHei）。"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        print("reportlab 未安装，跳过 PDF")
        return
    font = "SimHei"
    for cand in [r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\msyh.ttc"]:
        if Path(cand).exists():
            try:
                pdfmetrics.registerFont(TTFont(font, cand))
                break
            except Exception:
                continue
    doc = SimpleDocTemplate(str(out), pagesize=A4, leftMargin=14*mm,
                            rightMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm)
    style = ParagraphStyle("body", fontName=font, fontSize=9, leading=13)
    title = ParagraphStyle("title", fontName=font, fontSize=16, leading=22,
                           alignment=1, spaceAfter=8*mm)
    story = [Paragraph("需求调查问卷", title)]
    cur_chapter = None
    for q in questions:
        if q["chapter"] != cur_chapter:
            cur_chapter = q["chapter"]
            story.append(Paragraph(f"【{cur_chapter}】", style))
            story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f'{q["num"]}. [{q["importance"]}] {q["q"]}'
            + (f"（用途: {q['purpose']}）" if q["purpose"] else ""), style))
        story.append(Spacer(1, 4*mm))
    doc.build(story)
    print(f"PDF: {out}")


def gen_xlsx(questions: list[dict], out: Path):
    """用 openpyxl 生成审查表（3 Sheet：现有问题/遗漏补充/汇总）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl 未安装，跳过 XLSX")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "现有问题审查"
    ws.append(["序号", "章节", "问题", "必要度", "用途", "审查意见"])
    for q in questions:
        ws.append([q["num"], q["chapter"], q["q"], q["importance"],
                   q["purpose"], ""])
    ws2 = wb.create_sheet("遗漏问题补充")
    ws2.append(["补充问题", "说明"])
    ws3 = wb.create_sheet("优化建议汇总")
    ws3.append(["统计项", "数值"])
    ws3.append(["总问题数", len(questions)])
    core = sum(1 for q in questions if q["importance"].count("★") >= 4)
    ws3.append(["核心问题 (★★★★☆及以上)", core])
    ws3.append(["建议合并/删除", ""])
    # 高亮核心问题
    fill = PatternFill("solid", fgColor="FFFFE0")
    for row in ws.iter_rows(min_row=2):
        if row[3].value and row[3].value.count("★") >= 4:
            for c in row:
                c.fill = fill
    wb.save(str(out))
    print(f"XLSX: {out}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    md = Path(sys.argv[1])
    if not md.exists():
        print(f"问卷文件不存在: {md}")
        return 1
    questions = parse_questions(md)
    print(f"解析到 {len(questions)} 题")
    if not questions:
        print("未解析到题目（检查表格格式: | 序号 | 问题 | 必要度 | 用途 |）")
        return 1
    gen_pdf(questions, md.with_suffix(".pdf"))
    if "--pdf" not in sys.argv:
        gen_xlsx(questions, md.with_suffix(".xlsx"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
